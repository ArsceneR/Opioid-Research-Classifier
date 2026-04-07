"""
Evaluate classifier accuracy against human labels.

Compares classifier predictions (opioid_related vs neutral_content folders on Drive)
with human labels from the CSV. Reports full confusion matrix, precision, recall, F1,
and specificity.
"""

import pandas as pd
from pathlib import Path
import logging

logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')
logger = logging.getLogger(__name__)


def main():
    # excel_path = Path("/Users/arscenerubayita/Documents/Personal_Programming/Instagram_Scrape_and_Store/Coding Sheet_n=100(Sheet1).csv") #human labeled source of truth

    excel_path = Path("/Users/arscenerubayita/Documents/Personal_Programming/Instagram_Scrape_and_Store/Excel_With_IDs(Sheet1).csv") #human labeled source of truth
    opioid_dir = Path("/Users/arscenerubayita/Downloads/Opioid_Related_100_post_finetuning") #machine classified opioid_related posts
    neutral_dir = Path("/Users/arscenerubayita/Downloads/Neutral_Content_100_post_finetuning") #machine classified neutral posts

    # Load CSV with id and relevance columns
    df = pd.read_csv(excel_path)
    logger.info(f"CSV columns: {list(df.columns)}")

    id_col = None
    relevance_col = None
    for col in df.columns:
        col_lower = str(col).lower()
        if 'id' in col_lower and id_col is None:
            id_col = col
        if 'relevance' in col_lower and relevance_col is None:
            relevance_col = col

    if not id_col or not relevance_col:
        raise ValueError(f"Need 'id' and 'relevance' columns. Found: {list(df.columns)}")

    logger.info(f"Using ID column: '{id_col}', Relevance column: '{relevance_col}'")

    # Create mapping: id -> human relevance (1=opioid, 0=neutral)
    human_labels = {}
    for _, row in df.iterrows():
        post_id = str(row[id_col]).strip()
        relevance = int(row[relevance_col])
        human_labels[post_id] = relevance

    total_human_positive = sum(1 for v in human_labels.values() if v == 1)
    total_human_negative = sum(1 for v in human_labels.values() if v == 0)
    logger.info(f"Loaded {len(human_labels)} human labels (positive={total_human_positive}, negative={total_human_negative})")

    # Get classifier predictions from both folders
    clip_opioid_ids = set()
    clip_neutral_ids = set()

    if opioid_dir.exists():
        clip_opioid_ids = {d.name for d in opioid_dir.iterdir() if d.is_dir()}
        logger.info(f"Found {len(clip_opioid_ids)} folders in opioid_related")
    else:
        logger.warning(f"Opioid directory not found: {opioid_dir}")

    if neutral_dir.exists():
        clip_neutral_ids = {d.name for d in neutral_dir.iterdir() if d.is_dir()}
        logger.info(f"Found {len(clip_neutral_ids)} folders in neutral_content")
    else:
        logger.warning(f"Neutral directory not found: {neutral_dir}")

    if not clip_opioid_ids and not clip_neutral_ids:
        logger.error("No prediction folders found!")
        return

    # Build confusion matrix
    # TP: classifier=opioid, human=opioid (1)
    # FP: classifier=opioid, human=neutral (0)
    # TN: classifier=neutral, human=neutral (0)
    # FN: classifier=neutral, human=opioid (1)
    tp = fp = tn = fn = 0
    fp_ids = []
    fn_ids = []

    for post_id in clip_opioid_ids:
        if post_id in human_labels:
            if human_labels[post_id] == 1:
                tp += 1
            else:
                fp += 1
                fp_ids.append(post_id)

    for post_id in clip_neutral_ids:
        if post_id in human_labels:
            if human_labels[post_id] == 0:
                tn += 1
            else:
                fn += 1
                fn_ids.append(post_id)

    # Compute metrics
    precision = tp / (tp + fp) if (tp + fp) > 0 else 0.0
    recall = tp / (tp + fn) if (tp + fn) > 0 else 0.0
    f1 = 2 * precision * recall / (precision + recall) if (precision + recall) > 0 else 0.0
    specificity = tn / (tn + fp) if (tn + fp) > 0 else 0.0
    accuracy = (tp + tn) / (tp + fp + tn + fn) if (tp + fp + tn + fn) > 0 else 0.0

    # Print results to console
    print(f"\n{'='*50}")
    print(f"  CONFUSION MATRIX")
    print(f"{'='*50}")
    print(f"                    Predicted")
    print(f"                  Opioid  Neutral")
    print(f"  Actual Opioid   {tp:>5}    {fn:>5}")
    print(f"  Actual Neutral  {fp:>5}    {tn:>5}")
    print(f"{'='*50}")
    print(f"  METRICS")
    print(f"{'='*50}")
    print(f"  Accuracy:    {accuracy:.2%}  ({tp+tn}/{tp+fp+tn+fn})")
    print(f"  Precision:   {precision:.2%}  ({tp}/{tp+fp} predicted opioid are correct)")
    print(f"  Recall:      {recall:.2%}  ({tp}/{tp+fn} actual opioid detected)")
    print(f"  F1 Score:    {f1:.2%}")
    print(f"  Specificity: {specificity:.2%}  ({tn}/{tn+fp} actual neutral detected)")
    print(f"{'='*50}")

    if fp_ids:
        print(f"\n  False Positives (classifier=opioid, human=neutral):")
        for pid in sorted(fp_ids):
            print(f"    - {pid}")

    if fn_ids:
        print(f"\n  False Negatives (classifier=neutral, human=opioid):")
        for pid in sorted(fn_ids):
            print(f"    - {pid}")

    print()

    # Write results to Excel
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()

    # --- Sheet 1: Confusion Matrix & Metrics ---
    ws = wb.active
    ws.title = "Results"

    bold = Font(bold=True)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Confusion Matrix
    ws["A1"] = "Confusion Matrix"
    ws["A1"].font = Font(bold=True, size=14)

    # Column headers
    for col, label in [(3, "Predicted Opioid"), (4, "Predicted Neutral")]:
        cell = ws.cell(row=2, column=col, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Row: Actual Opioid
    ws.cell(row=3, column=2, value="Actual Opioid").font = bold
    ws.cell(row=3, column=3, value=tp).alignment = Alignment(horizontal="center")
    ws.cell(row=3, column=4, value=fn).alignment = Alignment(horizontal="center")

    # Row: Actual Neutral
    ws.cell(row=4, column=2, value="Actual Neutral").font = bold
    ws.cell(row=4, column=3, value=fp).alignment = Alignment(horizontal="center")
    ws.cell(row=4, column=4, value=tn).alignment = Alignment(horizontal="center")

    # Apply borders to matrix cells
    for row in range(2, 5):
        for col in range(2, 5):
            ws.cell(row=row, column=col).border = thin_border

    # Metrics table
    ws["A6"] = "Metrics"
    ws["A6"].font = Font(bold=True, size=14)

    metrics = [
        ("Accuracy", accuracy, f"{tp+tn}/{tp+fp+tn+fn}"),
        ("Precision", precision, f"{tp}/{tp+fp} predicted opioid are correct"),
        ("Recall (Sensitivity)", recall, f"{tp}/{tp+fn} actual opioid detected"),
        ("F1 Score", f1, ""),
        ("Specificity", specificity, f"{tn}/{tn+fp} actual neutral detected"),
    ]

    for col, header in [(1, "Metric"), (2, "Value"), (3, "Detail")]:
        cell = ws.cell(row=7, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for i, (name, value, detail) in enumerate(metrics, start=8):
        ws.cell(row=i, column=1, value=name).border = thin_border
        val_cell = ws.cell(row=i, column=2, value=value)
        val_cell.number_format = "0.00%"
        val_cell.border = thin_border
        ws.cell(row=i, column=3, value=detail).border = thin_border

    # Auto-fit column widths
    for col_cells in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col_cells), default=0)
        ws.column_dimensions[col_cells[0].column_letter].width = max_len + 4

    # --- Sheet 2: False Positives ---
    ws_fp = wb.create_sheet("False Positives")
    ws_fp.cell(row=1, column=1, value="Post ID").font = bold
    ws_fp.cell(row=1, column=2, value="Classifier").font = bold
    ws_fp.cell(row=1, column=3, value="Human Label").font = bold
    for i, pid in enumerate(sorted(fp_ids), start=2):
        ws_fp.cell(row=i, column=1, value=pid)
        ws_fp.cell(row=i, column=2, value="opioid")
        ws_fp.cell(row=i, column=3, value="neutral")

    # --- Sheet 3: False Negatives ---
    ws_fn = wb.create_sheet("False Negatives")
    ws_fn.cell(row=1, column=1, value="Post ID").font = bold
    ws_fn.cell(row=1, column=2, value="Classifier").font = bold
    ws_fn.cell(row=1, column=3, value="Human Label").font = bold
    for i, pid in enumerate(sorted(fn_ids), start=2):
        ws_fn.cell(row=i, column=1, value=pid)
        ws_fn.cell(row=i, column=2, value="neutral")
        ws_fn.cell(row=i, column=3, value="opioid")

    output_xlsx = excel_path.parent / "classifier_evaluation.xlsx"
    wb.save(output_xlsx)
    logger.info(f"Results saved to {output_xlsx}")


if __name__ == '__main__':
    main()
